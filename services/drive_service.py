import io
import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime
from PIL import Image

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2.service_account import Credentials

from config.settings import settings
from services.json_store import cargar_json, guardar_json, obtener_lock, user_data_path
from services.timezone_utils import ahora

logger = logging.getLogger(__name__)

# ---------- Google Drive API singleton ----------

_drive_client = None
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def _get_drive_service():
    global _drive_client
    if _drive_client is None:
        creds = Credentials.from_service_account_file(
            settings.GOOGLE_SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        _drive_client = build("drive", "v3", credentials=creds, cache_discovery=False)
    return _drive_client


# ---------- Staging temporal por usuario ----------

_staging: dict[str, dict] = {}


# ---------- Helpers internos ----------

def _registro_path(user_id: str) -> Path:
    return user_data_path(user_id, "registro.json")


def _capturas_dir(user_id: str) -> Path:
    return user_data_path(user_id, "capturas")


def _cargar_registro(user_id: str) -> list[dict]:
    return cargar_json(_registro_path(user_id))


def _guardar_registro(user_id: str, registro: list[dict]):
    guardar_json(_registro_path(user_id), registro)


def _extraer_id_drive(url: str) -> str | None:
    """Extrae el ID de archivo o carpeta de un link de Google Drive."""
    patrones = [
        r"/folders/([a-zA-Z0-9_-]+)",
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"id=([a-zA-Z0-9_-]+)",
        r"/d/([a-zA-Z0-9_-]+)",
    ]
    for patron in patrones:
        match = re.search(patron, url)
        if match:
            return match.group(1)
    return None


# ---------- Operaciones de bytes en memoria ----------

def _descargar_bytes(file_id: str) -> bytes:
    """Descarga un archivo de Drive a bytes en memoria."""
    service = _get_drive_service()
    request = service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def _exportar_bytes(file_id: str, mime_type: str) -> bytes:
    """Exporta un archivo nativo de Google (Docs/Sheets) a bytes."""
    service = _get_drive_service()
    request = service.files().export_media(fileId=file_id, mimeType=mime_type)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def _detectar_tipo(nombre: str, mime_type: str) -> str:
    """Detecta el tipo de archivo por extension o MIME type."""
    ext = Path(nombre).suffix.lower() if nombre else ""
    if ext == ".pdf" or mime_type == "application/pdf":
        return "pdf"
    if ext in (".xlsx", ".xls") or "spreadsheet" in mime_type:
        return "xlsx"
    if mime_type == "application/vnd.google-apps.spreadsheet":
        return "google_sheet"
    if mime_type == "application/vnd.google-apps.document":
        return "google_doc"
    return "texto"


def _es_carpeta(mime_type: str) -> bool:
    return mime_type == "application/vnd.google-apps.folder"


def _obtener_bytes_archivo(file_id: str, mime_type: str) -> tuple[bytes, str]:
    """Obtiene bytes de un archivo, exportando si es nativo de Google.

    Returns (data, tipo_efectivo). Raises ValueError si es carpeta.
    """
    if _es_carpeta(mime_type):
        raise ValueError("Es una carpeta, no un archivo descargable")
    if mime_type == "application/vnd.google-apps.spreadsheet":
        data = _exportar_bytes(file_id, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return data, "xlsx"
    if mime_type == "application/vnd.google-apps.document":
        data = _exportar_bytes(file_id, "application/pdf")
        return data, "pdf"
    data = _descargar_bytes(file_id)
    return data, _detectar_tipo("", mime_type)


def _listar_archivos_recursivo(service, folder_id: str, max_total: int = 50) -> list[dict]:
    """Lista archivos dentro de una carpeta, entrando recursivamente en subcarpetas."""
    resultado = []

    resp = service.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id,name,mimeType)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        pageSize=100,
    ).execute()

    for f in resp.get("files", []):
        if len(resultado) >= max_total:
            break
        if _es_carpeta(f["mimeType"]):
            sub = _listar_archivos_recursivo(service, f["id"], max_total - len(resultado))
            # Prefijar con nombre de subcarpeta para claridad
            for sf in sub:
                sf["name"] = f"{f['name']}/{sf['name']}"
            resultado.extend(sub)
        else:
            resultado.append(f)

    return resultado


def _parsear_pdf(data: bytes, max_paginas: int = 10) -> dict:
    """Lee PDF desde bytes."""
    import fitz
    doc = fitz.open(stream=data, filetype="pdf")
    total_pags = len(doc)
    paginas = []
    for i, pagina in enumerate(doc):
        if i >= max_paginas:
            break
        paginas.append(f"--- Pagina {i+1} ---\n{pagina.get_text()}")
    doc.close()
    return {
        "contenido": "\n\n".join(paginas),
        "total_paginas": total_pags,
    }


def _parsear_xlsx(data: bytes, max_lineas: int = 100) -> dict:
    """Lee XLSX desde bytes."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    resultado = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        resultado.append(f"--- Hoja: {nombre_hoja} ---")
        lineas_hoja = 0
        for row in ws.iter_rows(values_only=True):
            if lineas_hoja >= max_lineas:
                resultado.append(f"... (truncado a {max_lineas} lineas)")
                break
            celdas = [str(c) if c is not None else "" for c in row]
            resultado.append(" | ".join(celdas))
            lineas_hoja += 1
    wb.close()
    return {"contenido": "\n".join(resultado)}


def _parsear_texto(data: bytes, max_lineas: int = 100) -> dict:
    """Lee archivo de texto desde bytes."""
    texto = data.decode("utf-8", errors="replace")
    lineas = texto.splitlines()
    return {
        "contenido": "\n".join(lineas[:max_lineas]),
        "total_lineas": len(lineas),
        "truncado": len(lineas) > max_lineas,
    }


# ---------- Migracion de registros existentes ----------

def _migrar_registro_v2(user_id: str):
    """Migra entradas existentes: extrae drive_id de url_drive, borra campo ruta."""
    path = _registro_path(user_id)
    lock = obtener_lock(path)
    with lock:
        registro = cargar_json(path)
        if not registro:
            return
        modificado = False
        for entrada in registro:
            if "ruta" in entrada and "drive_id" not in entrada:
                drive_id = _extraer_id_drive(entrada.get("url_drive", ""))
                if drive_id:
                    entrada["drive_id"] = drive_id
                del entrada["ruta"]
                modificado = True
        if modificado:
            guardar_json(path, registro)
            logger.info("Registro migrado a v2 para user %s", user_id)


# ---------- Funciones principales ----------

def conectar_drive(user_id: str, url: str, nombre: str | None = None) -> dict:
    """Conecta un archivo o carpeta de Google Drive via API (sin descargar)."""
    drive_id = _extraer_id_drive(url)
    if not drive_id:
        return {"ok": False, "error": "No se pudo extraer el ID del link de Drive."}

    es_carpeta = "/folders/" in url

    try:
        service = _get_drive_service()

        if es_carpeta:
            # Obtener metadata de la carpeta
            meta = service.files().get(
                fileId=drive_id, fields="id,name", supportsAllDrives=True
            ).execute()
            # Listar archivos dentro
            resp = service.files().list(
                q=f"'{drive_id}' in parents and trashed=false",
                fields="files(id,name,mimeType)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                pageSize=100,
            ).execute()
            archivos_drive = resp.get("files", [])
            archivos_nombres = [f["name"] for f in archivos_drive]
            archivos_ids = {f["name"]: {"id": f["id"], "mimeType": f["mimeType"]} for f in archivos_drive}
            nombre_sugerido = nombre or meta.get("name", drive_id)

            # Guardar en staging (con drive_id, NUNCA expuesto al LLM)
            _staging[user_id] = {
                "drive_id": drive_id,
                "tipo": "carpeta",
                "nombre": nombre_sugerido,
                "archivos": archivos_nombres,
                "archivos_ids": archivos_ids,
                "url": url,
            }

            return {
                "ok": True,
                "archivos": archivos_nombres,
                "tipo": "carpeta",
                "nombre_sugerido": nombre_sugerido,
            }
        else:
            # Archivo individual
            meta = service.files().get(
                fileId=drive_id, fields="id,name,mimeType", supportsAllDrives=True
            ).execute()
            nombre_archivo = meta.get("name", drive_id)
            nombre_sugerido = nombre or nombre_archivo

            _staging[user_id] = {
                "drive_id": drive_id,
                "tipo": "archivo",
                "nombre": nombre_sugerido,
                "archivos": [nombre_archivo],
                "mime_type": meta.get("mimeType", ""),
                "url": url,
            }

            return {
                "ok": True,
                "archivos": [nombre_archivo],
                "tipo": "archivo",
                "nombre_sugerido": nombre_sugerido,
            }

    except Exception as e:
        error_str = str(e)
        if "404" in error_str or "notFound" in error_str:
            return {"ok": False, "error": (
                "Error 404: Archivo no encontrado en Drive. "
                "Puede que el link este mal, el archivo fue eliminado, o no esta compartido. "
                "El usuario debe compartirlo con syn-drive-bot@syn-cortex-489502.iam.gserviceaccount.com"
            )}
        if "403" in error_str or "forbidden" in error_str:
            return {"ok": False, "error": (
                "Error 403: Sin permisos de acceso. "
                "El usuario debe compartir el archivo/carpeta con syn-drive-bot@syn-cortex-489502.iam.gserviceaccount.com "
                "con permisos de Lector. Si ya lo hizo, esperar unos minutos y reintentar."
            )}
        return {"ok": False, "error": f"Error al conectar con Drive: {e}"}


def agregar_al_registro(
    user_id: str, nombre: str, descripcion: str,
    url_drive: str, tipo: str, archivos: list[str],
) -> str:
    """Agrega una entrada al registro usando datos del staging."""
    staging = _staging.get(user_id)
    if not staging:
        return "Error: no hay conexion pendiente. Primero usa conectar_drive."

    drive_id = staging["drive_id"]

    path = _registro_path(user_id)
    lock = obtener_lock(path)
    with lock:
        registro = cargar_json(path)
        for entrada in registro:
            if entrada.get("url_drive") == url_drive:
                entrada["descripcion"] = descripcion
                entrada["nombre"] = nombre
                entrada["drive_id"] = drive_id
                if tipo == "carpeta":
                    entrada["archivos_ids"] = staging.get("archivos_ids", {})
                guardar_json(path, registro)
                _staging.pop(user_id, None)
                return f"Registro actualizado: {nombre}"

        entrada_nueva = {
            "nombre": nombre,
            "descripcion": descripcion,
            "url_drive": url_drive,
            "tipo": tipo,
            "archivos": archivos,
            "drive_id": drive_id,
        }
        if tipo == "carpeta":
            entrada_nueva["archivos_ids"] = staging.get("archivos_ids", {})
        registro.append(entrada_nueva)
        guardar_json(path, registro)

    _staging.pop(user_id, None)
    return f"Registrado: {nombre} ({len(archivos)} archivo(s))"


def listar_registro(user_id: str) -> str:
    """Devuelve el contenido del registro con info dinamica de Drive."""
    _migrar_registro_v2(user_id)
    registro = _cargar_registro(user_id)
    if not registro:
        return "El registro esta vacio. No hay archivos guardados."

    lineas = []
    for i, entrada in enumerate(registro, 1):
        drive_id = entrada.get("drive_id", "")
        info_extra = ""

        try:
            service = _get_drive_service()
            if entrada["tipo"] == "carpeta" and drive_id:
                archivos_actuales = _listar_archivos_recursivo(service, drive_id)
                nombres = [f["name"] for f in archivos_actuales]
                info_extra = (
                    f"   Archivos actuales: {len(nombres)} — "
                    f"{', '.join(nombres[:10])}"
                    f"{'...' if len(nombres) > 10 else ''}"
                )
            elif drive_id:
                meta = service.files().get(
                    fileId=drive_id,
                    fields="modifiedTime",
                    supportsAllDrives=True,
                ).execute()
                mod_time = meta.get("modifiedTime", "desconocida")
                if mod_time and mod_time != "desconocida":
                    mod_time = mod_time[:19].replace("T", " ")
                info_extra = f"   Ultima modificacion: {mod_time}"
        except Exception:
            info_extra = "   (no se pudo consultar info de Drive)"

        bloque = (
            f"{i}. {entrada['nombre']} [{entrada['tipo']}]\n"
            f"   Descripcion: {entrada['descripcion']}\n"
            f"   URL: {entrada.get('url_drive', 'N/A')}"
        )
        if info_extra:
            bloque += f"\n{info_extra}"
        lineas.append(bloque)

    return "\n\n".join(lineas)


def buscar_en_registro(user_id: str, termino: str) -> str:
    """Busca archivos en el registro por nombre o descripcion."""
    registro = _cargar_registro(user_id)
    termino_lower = termino.lower()
    resultados = []

    for entrada in registro:
        if (
            termino_lower in entrada["nombre"].lower()
            or termino_lower in entrada["descripcion"].lower()
            or any(termino_lower in a.lower() for a in entrada.get("archivos", []))
        ):
            resultados.append(
                f"- {entrada['nombre']}: {entrada['descripcion'][:200]}"
            )

    if not resultados:
        return f"No se encontraron archivos relacionados con '{termino}'."
    return "Resultados:\n" + "\n".join(resultados)


def editar_descripcion(user_id: str, nombre: str, nueva_descripcion: str) -> str:
    """Edita la descripcion de un archivo en el registro."""
    path = _registro_path(user_id)
    lock = obtener_lock(path)
    with lock:
        registro = cargar_json(path)
        for entrada in registro:
            if entrada["nombre"].lower() == nombre.lower():
                entrada["descripcion"] = nueva_descripcion
                guardar_json(path, registro)
                return f"Descripcion de '{nombre}' actualizada."
    return f"No se encontro '{nombre}' en el registro."


def eliminar_del_registro(user_id: str, nombre: str) -> str:
    """Elimina una entrada del registro."""
    path = _registro_path(user_id)
    lock = obtener_lock(path)
    with lock:
        registro = cargar_json(path)
        nuevo = [e for e in registro if e["nombre"].lower() != nombre.lower()]
        if len(nuevo) == len(registro):
            return f"No se encontro '{nombre}' en el registro."
        guardar_json(path, nuevo)
    return f"'{nombre}' eliminado del registro."


def leer_archivo(user_id: str, nombre: str, max_lineas: int = 100) -> dict:
    """Lee el contenido de un archivo registrado directamente desde Drive."""
    _migrar_registro_v2(user_id)
    registro = _cargar_registro(user_id)

    entrada = None
    for e in registro:
        if e["nombre"].lower() == nombre.lower():
            entrada = e
            break

    if not entrada:
        return {"ok": False, "error": f"No se encontro '{nombre}' en el registro."}

    drive_id = entrada.get("drive_id")
    if not drive_id:
        return {"ok": False, "error": "Este archivo no tiene drive_id. Vuelve a conectarlo."}

    try:
        service = _get_drive_service()

        if entrada["tipo"] == "carpeta":
            # Listar archivos recursivamente (entra en subcarpetas)
            archivos_drive = _listar_archivos_recursivo(service, drive_id)

            archivos_contenido = {}
            errores = []
            for arch in archivos_drive:
                try:
                    data, tipo_eff = _obtener_bytes_archivo(
                        arch["id"], arch.get("mimeType", "")
                    )
                    if len(data) > 2_000_000:
                        archivos_contenido[arch["name"]] = {"contenido": "[Archivo muy grande (>2MB), omitido]"}
                        continue

                    if tipo_eff == "pdf":
                        parsed = _parsear_pdf(data)
                    elif tipo_eff == "xlsx":
                        parsed = _parsear_xlsx(data, max_lineas)
                    else:
                        parsed = _parsear_texto(data, max_lineas)
                    archivos_contenido[arch["name"]] = parsed
                except Exception as ex:
                    msg_error = f"Error leyendo '{arch['name']}': {ex}"
                    errores.append(msg_error)
                    archivos_contenido[arch["name"]] = {"contenido": f"[{msg_error}]"}

            resultado = {
                "ok": True,
                "nombre": entrada["nombre"],
                "tipo": "carpeta",
                "archivos": archivos_contenido,
                "total_archivos": len(archivos_drive),
                "archivos_leidos": len(archivos_drive) - len(errores),
                "descripcion": entrada["descripcion"],
            }
            if errores:
                resultado["errores"] = errores
                resultado["nota_errores"] = (
                    f"ATENCION: {len(errores)} archivo(s) no se pudieron leer. "
                    "Muestra estos errores al usuario para que pueda reportarlos."
                )
            return resultado

        # Archivo individual
        meta = service.files().get(
            fileId=drive_id, fields="name,mimeType", supportsAllDrives=True
        ).execute()
        mime_type = meta.get("mimeType", "")

        data, tipo_eff = _obtener_bytes_archivo(drive_id, mime_type)

        if tipo_eff == "pdf":
            parsed = _parsear_pdf(data)
            return {
                "ok": True,
                "nombre": entrada["nombre"],
                "tipo": "pdf",
                "descripcion": entrada["descripcion"],
                **parsed,
            }
        elif tipo_eff == "xlsx":
            parsed = _parsear_xlsx(data, max_lineas)
            return {
                "ok": True,
                "nombre": entrada["nombre"],
                "tipo": "xlsx",
                "descripcion": entrada["descripcion"],
                **parsed,
            }
        else:
            parsed = _parsear_texto(data, max_lineas)
            return {
                "ok": True,
                "nombre": entrada["nombre"],
                "tipo": "texto",
                "descripcion": entrada["descripcion"],
                **parsed,
            }

    except Exception as e:
        error_str = str(e)
        if "404" in error_str or "notFound" in error_str:
            return {"ok": False, "error": "Archivo no encontrado en Drive. Puede haber sido eliminado o ya no esta compartido."}
        if "403" in error_str or "forbidden" in error_str:
            return {"ok": False, "error": "Sin permisos para leer. Verifica que el archivo siga compartido con el service account."}
        return {"ok": False, "error": f"Error al leer desde Drive: {e}"}


# ---------- Captura de citas ----------

def _buscar_archivo_en_carpeta(service, folder_drive_id: str, nombre_archivo: str) -> dict | None:
    """Busca un archivo por nombre dentro de una carpeta (recursivo)."""
    archivos = _listar_archivos_recursivo(service, folder_drive_id)
    nombre_lower = nombre_archivo.lower()
    # Coincidencia exacta
    for a in archivos:
        base = a["name"].rsplit("/", 1)[-1] if "/" in a["name"] else a["name"]
        if base.lower() == nombre_lower or a["name"].lower() == nombre_lower:
            return a
    # Coincidencia parcial
    for a in archivos:
        base = a["name"].rsplit("/", 1)[-1] if "/" in a["name"] else a["name"]
        if nombre_lower in base.lower() or nombre_lower in a["name"].lower():
            return a
    return None


def _renderizar_archivo(data: bytes, filetype: str, texto_cita: str = "") -> Image.Image:
    """Renderiza cualquier archivo soportado por PyMuPDF como imagen real.

    PyMuPDF soporta PDF, XLSX, DOCX, entre otros.
    Para archivos multi-pagina, busca la pagina con el texto citado.
    """
    import fitz
    doc = fitz.open(stream=data, filetype=filetype)

    # Buscar la pagina que mejor coincide con el texto citado
    mejor_pagina = 0
    if texto_cita and len(doc) > 1:
        mejor_score = 0
        palabras_cita = set(texto_cita.lower().split()[:20])
        for i, pagina in enumerate(doc):
            texto_pag = pagina.get_text().lower()
            coincidencias = sum(1 for p in palabras_cita if p in texto_pag)
            if coincidencias > mejor_score:
                mejor_score = coincidencias
                mejor_pagina = i

    pagina = doc[mejor_pagina]
    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom para buena calidad
    pix = pagina.get_pixmap(matrix=mat)
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    doc.close()
    return img


def _convertir_a_pdf_con_libreoffice(data: bytes, extension: str) -> bytes | None:
    """Convierte un archivo (XLSX, DOCX, etc.) a PDF usando LibreOffice headless.

    Retorna los bytes del PDF generado, o None si LibreOffice no esta disponible.
    """
    import subprocess
    import tempfile
    import shutil

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / f"input{extension}"
        input_path.write_bytes(data)

        try:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, str(input_path)],
                capture_output=True, timeout=30,
            )
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return None

        pdf_path = Path(tmpdir) / "input.pdf"
        if pdf_path.exists():
            return pdf_path.read_bytes()
    return None


def _es_pdf(nombre: str, mime_type: str) -> bool:
    ext = Path(nombre).suffix.lower() if nombre else ""
    return ext == ".pdf" or mime_type == "application/pdf"


def _extension_archivo(nombre: str, mime_type: str) -> str:
    ext = Path(nombre).suffix.lower() if nombre else ""
    if ext:
        return ext
    if "spreadsheet" in mime_type:
        return ".xlsx"
    if "document" in mime_type:
        return ".docx"
    return ".bin"


def generar_captura(
    user_id: str,
    nombre_archivo: str,
    texto_cita: str,
    contexto: str = "",
) -> dict:
    """Genera una captura renderizando el archivo real desde Drive con PyMuPDF."""
    capturas_dir = _capturas_dir(user_id)
    capturas_dir.mkdir(parents=True, exist_ok=True)

    nombre_sanitizado = re.sub(r'[^\w.-]', '_', nombre_archivo)
    ts = ahora().strftime("%Y%m%d_%H%M%S_%f")
    nombre_img = f"cita_{nombre_sanitizado}_{ts}.png"
    ruta_img = capturas_dir / nombre_img

    try:
        registro = _cargar_registro(user_id)
        service = _get_drive_service()
        data = None
        file_name = nombre_archivo
        mime_type = ""

        # Buscar el archivo en el registro
        for entrada in registro:
            if entrada["tipo"] == "carpeta":
                drive_id = entrada.get("drive_id", "")
                if not drive_id:
                    continue
                arch = _buscar_archivo_en_carpeta(service, drive_id, nombre_archivo)
                if arch:
                    data, _ = _obtener_bytes_archivo(arch["id"], arch.get("mimeType", ""))
                    file_name = arch["name"].rsplit("/", 1)[-1] if "/" in arch["name"] else arch["name"]
                    mime_type = arch.get("mimeType", "")
                    break
            elif entrada["nombre"].lower() == nombre_archivo.lower():
                drive_id = entrada.get("drive_id", "")
                if not drive_id:
                    continue
                meta = service.files().get(
                    fileId=drive_id, fields="name,mimeType", supportsAllDrives=True
                ).execute()
                data, _ = _obtener_bytes_archivo(drive_id, meta.get("mimeType", ""))
                file_name = meta.get("name", nombre_archivo)
                mime_type = meta.get("mimeType", "")
                break

        if not data:
            return {"ok": False, "error": f"No se encontro '{nombre_archivo}' para generar captura."}

        if _es_pdf(file_name, mime_type):
            # PDF: renderizar directamente con PyMuPDF
            img = _renderizar_archivo(data, "pdf", texto_cita)
        else:
            # XLSX, DOCX, etc: convertir a PDF con LibreOffice, luego renderizar
            ext = _extension_archivo(file_name, mime_type)
            pdf_data = _convertir_a_pdf_con_libreoffice(data, ext)
            if not pdf_data:
                return {
                    "ok": False,
                    "error": (
                        f"No se pudo generar captura de '{nombre_archivo}'. "
                        "LibreOffice no esta disponible para convertir el archivo. "
                        "Los datos citados en el mensaje vienen directamente del archivo original."
                    ),
                }
            img = _renderizar_archivo(pdf_data, "pdf", texto_cita)

        img.save(str(ruta_img), "PNG")

        return {
            "ok": True,
            "ruta_imagen": str(ruta_img),
            "nombre_archivo": nombre_archivo,
            "ancho": img.width,
            "alto": img.height,
        }

    except Exception as e:
        logger.warning("Error generando captura de %s: %s", nombre_archivo, e)
        return {"ok": False, "error": f"No se pudo generar captura de '{nombre_archivo}': {e}"}
